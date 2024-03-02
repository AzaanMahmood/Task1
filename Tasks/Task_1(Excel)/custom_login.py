from openpyxl import load_workbook

def read_data(file_name):
    workbook = load_workbook(file_name)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
        print("{:<15} {:<10} {:<10} {:<15}".format(*row))
    return data

def add_new_data(data):
    name = input("Enter Name: ")
    age = input("Enter Age: ")  # Keeping age as a string
    gender = input("Enter Gender: ")
    city = input("Enter City: ")

    new_row = [name, age, gender, city]
    data.append(new_row)
    return data

def filter_data(data):
    age_filter = input("Enter the age filter: ")  # Keeping age filter as a string
    filtered_data = [row for row in data if str(row[1]) == age_filter]
    if filtered_data:
        print("\nFiltered Data:")
        print("{:<15} {:<10} {:<10} {:<15}".format("Name", "Age", "Gender", "City"))
        for row in filtered_data:
            print("{:<15} {:<10} {:<10} {:<15}".format(*row))
    else:
        print("No data matching the filter criteria.")

data = read_data('custom.xlsx')

if input("Do you want to add new data? (yes/no): ").lower() == 'yes':
    data = add_new_data(data)
    print("New data has been added.")

if input("Do you want to filter data? (yes/no): ").lower() == 'yes':
    filter_data(data)
